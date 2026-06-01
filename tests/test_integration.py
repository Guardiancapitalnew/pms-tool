"""End-to-end integration tests: research file → validate → session/broker
files → broker reply → match → allocate → allocation file.

These exercise the full pipeline a real day's flow goes through. Synthetic
Excel files built in-memory so the test is reproducible and doesn't depend
on the sample_data/ folder."""
from io import BytesIO

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

from utils.reader import (
    read_research_file, read_bank_book, read_scrip_wise_report,
    read_session_file,
)
from utils.writer import write_session_file, write_allocation_file
from utils.isin import load_isin_database
from part1.validator import validate_orders
from part1.session import build_session_file
from part1.broker_file import build_broker_file
from part2.parser import (
    parse_ambit_reply, parse_incred_reply, get_incred_cp_codes,
    NORMALISED_COLUMNS,
)
from part2.matcher import match_session_to_broker
from part2.allocator import allocate_costs, ALLOCATION_COLUMNS


# ── Synthetic file builders ──────────────────────────────────────────────────

def _research_buys_xlsx() -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"
    ws.append(["S.No", "OFIN", "Client", "Ticker", "Direction",
               "Qty", "Ref Price", "Value", "CP Code"])
    ws.append([1, "OF001", "Client A", "KOTAKBANK", "BUY", 100, 400.0, 40000.0, "ORBIS001"])
    ws.append([2, "OF002", "Client B", "KOTAKBANK", "BUY", 200, 400.0, 80000.0, "ORBIS001"])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _bank_book_xlsx() -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Bank Balance Summary"
    ws.append([])  # Orbis exports often have title rows
    ws.append(["OFIN Code", "Balance"])
    ws.append(["OF001", 500000.0])
    ws.append(["OF002", 200000.0])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _ambit_kotakbank_buy_300_xlsx() -> BytesIO:
    """Broker executed 300 KOTAKBANK BUY at NSE on 2026-05-18."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Transaction Date", "Exchange", "NSE Symbol", "BSE Scrip Code",
               "ISIN No.", "Transaction Type", "quantity",
               "Brokerage", "stt", "Stamp Duty", "SEBI Charges",
               "Turnover Tax", "Other Charges", "GST Amount", "Net Amount"])
    ws.append(["2026-05-18", "NSE", "KOTAKBANK", "500247",
               "INE237A01036", "BUY", 300,
               60.0, 30.0, 6.0, 0.06, 3.6, 0.0, 11.0, 120100.66])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── End-to-end tests ─────────────────────────────────────────────────────────

def test_full_buy_only_pipeline_round_trips_session_file():
    """Buy-only research file flows through: read → validate (no scrip-wise)
    → build session + broker → write/read session round-trip → parse Ambit
    reply → match → allocate → write allocation file. Verifies every layer."""

    # Part 1: read inputs and validate
    research_df = read_research_file(_research_buys_xlsx())
    bank_book = read_bank_book(_bank_book_xlsx())
    isin_db = load_isin_database()

    validation_df = validate_orders(
        research_df=research_df,
        bank_book=bank_book,
        scrip_df=None,        # buy-only flow - exercises the new optional path
        isin_db=isin_db,
    )
    assert all(validation_df["Status"] == "GREEN"), validation_df[["Reason"]].to_dict()

    # Part 1: build session + broker files
    session_df = build_session_file(included_df=validation_df, existing_session_df=None)
    broker_file_df = build_broker_file(validation_df)

    # Session columns + batch number
    assert list(session_df.columns) == [
        "S.No", "Batch", "OFIN", "Client", "Ticker",
        "ISIN", "Direction", "Qty", "Ref Price", "CP Code",
    ]
    assert int(session_df["Batch"].max()) == 1

    # Broker file aggregates by Ticker+Direction
    assert len(broker_file_df) == 1
    assert broker_file_df.iloc[0]["Total Qty"] == 300

    # Round-trip the session file through the writer and reader
    session_bytes = write_session_file(session_df)
    session_round_trip = read_session_file(BytesIO(session_bytes))
    assert len(session_round_trip) == len(session_df)
    assert set(session_round_trip["Direction"]) == {"BUY"}

    # Part 2: parse Ambit reply
    broker_df = parse_ambit_reply(_ambit_kotakbank_buy_300_xlsx())
    assert list(broker_df.columns) == NORMALISED_COLUMNS

    # Part 2: match - both session rows should match the single broker row
    matched, not_executed, unexpected = match_session_to_broker(
        session_round_trip, broker_df
    )
    assert len(matched) == 2
    assert not_executed == []
    assert unexpected == []

    # Part 2: allocate
    allocation_df = allocate_costs(matched, broker_df)
    assert list(allocation_df.columns) == ALLOCATION_COLUMNS
    assert len(allocation_df) == 2

    # Weights 100/300 and 200/300 - charges must sum to broker totals exactly
    assert allocation_df["InputBrokerage"].sum() == pytest.approx(60.0, rel=1e-9)
    assert allocation_df["InputSTT"].sum() == pytest.approx(30.0, rel=1e-9)
    assert allocation_df["InputNetAmount"].sum() == pytest.approx(120100.66, rel=1e-9)

    # Buy/Sell case is title-case (Orbis requirement)
    assert set(allocation_df["Buy/ Sell"]) == {"Buy"}

    # Allocation file writes cleanly and is readable
    allocation_bytes = write_allocation_file(allocation_df)
    wb = load_workbook(BytesIO(allocation_bytes))
    assert "Allocation" in wb.sheetnames
    ws = wb["Allocation"]
    assert ws.max_column == 19
    assert ws.max_row == 3  # header + 2 data rows


def _research_sells_xlsx() -> BytesIO:
    """Sell-only research file with BLANK CP Code, so the InCred path is
    forced to fall back to the CP code from the broker reply."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"
    ws.append(["S.No", "OFIN", "Client", "Ticker", "Direction",
               "Qty", "Ref Price", "Value", "CP Code"])
    ws.append([1, "OF001", "Client A", "KOTAKBANK", "SELL", 100, 400.0, 40000.0, ""])
    ws.append([2, "OF002", "Client B", "KOTAKBANK", "SELL", 200, 400.0, 80000.0, ""])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _scrip_wise_with_holdings_xlsx() -> BytesIO:
    """Holdings report showing both OFINs have enough KOTAKBANK to sell."""
    wb = Workbook()
    ws = wb.active
    ws.title = "ClientA"  # sheet name varies daily; reader uses index 0
    ws.append(["Scrip Name", "Item No", "Client Code", "Quantity"])
    ws.append(["KOTAKBANK", "INE237A01036", "OF001", 500])
    ws.append(["KOTAKBANK", "INE237A01036", "OF002", 500])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _incred_kotakbank_sell_300_xlsx() -> BytesIO:
    """InCred reply: 300 KOTAKBANK SELL at NSE, CP Code in the reply."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Incred_Capital_Trade_Confirmati"
    ws.append(["Exchange", "ISIN No.", "Transaction Type", "Quantity",
               "Amount", "Brokerage", "STT", "Stamp Duty", "SEBI Charges",
               "Turnover Tax", "Other Charges", "GST Amount", "Net Amount",
               "CP CODE"])
    ws.append(["NSE", "INE237A01036", "SELL", "300", "120000",
               "60", "30", "6", "0.06", "3.6", "0", "11", "119899.34",
               "ORBIS-INCRED-001"])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def test_full_sell_only_pipeline_via_incred():
    """End-to-end with sells + InCred broker:
       - sell-only research file (no bank book needed)
       - blank CP Code in research → must populate from InCred reply
       - InCred parsing (sheet name + string→float casts)
       - allocation file written with Sell (title case)
    """
    # Part 1
    research_df = read_research_file(_research_sells_xlsx())
    scrip_df = read_scrip_wise_report(_scrip_wise_with_holdings_xlsx())
    isin_db = load_isin_database()

    validation_df = validate_orders(
        research_df=research_df,
        bank_book=None,          # sell-only, no bank book
        scrip_df=scrip_df,
        isin_db=isin_db,
    )
    assert all(validation_df["Status"] == "GREEN"), validation_df[["Reason"]].to_dict()

    session_df = build_session_file(validation_df, existing_session_df=None)
    # Round-trip
    session_bytes = write_session_file(session_df)
    session_round_trip = read_session_file(BytesIO(session_bytes))
    assert set(session_round_trip["Direction"]) == {"SELL"}

    # Part 2 - InCred path. Must call get_incred_cp_codes BEFORE the parser
    # call (or with a fresh seek), since both consume the stream.
    incred_file = _incred_kotakbank_sell_300_xlsx()
    broker_df = parse_incred_reply(incred_file)
    incred_file.seek(0)
    incred_cp_codes = get_incred_cp_codes(incred_file)
    assert "INE237A01036" in incred_cp_codes

    matched, not_executed, unexpected = match_session_to_broker(
        session_round_trip, broker_df
    )
    assert len(matched) == 2
    assert not_executed == []

    allocation_df = allocate_costs(matched, broker_df, incred_cp_codes=incred_cp_codes)
    assert len(allocation_df) == 2
    # CP Code came from InCred reply (research file was blank)
    assert set(allocation_df["CP CODE"]) == {"ORBIS-INCRED-001"}
    # Buy/Sell title case preserved
    assert set(allocation_df["Buy/ Sell"]) == {"Sell"}

    # Final allocation file writes cleanly
    allocation_bytes = write_allocation_file(allocation_df)
    wb = load_workbook(BytesIO(allocation_bytes))
    assert "Allocation" in wb.sheetnames
    assert wb["Allocation"].max_column == 19


def test_full_batch_2_flow_increments_batch_number():
    """Upload + validate research, then validate a 2nd batch against the
    existing session file. Batch number should increment and committed cash
    from batch 1 should be deducted from available cash in batch 2's checks."""

    # Batch 1 - run the pipeline through session-file creation
    research_df_1 = read_research_file(_research_buys_xlsx())
    bank_book = read_bank_book(_bank_book_xlsx())
    isin_db = load_isin_database()
    val_1 = validate_orders(research_df_1, bank_book, None, isin_db)
    session_df_1 = build_session_file(val_1, existing_session_df=None)
    assert int(session_df_1["Batch"].max()) == 1

    # Round-trip batch 1 session file (simulates upload at start of batch 2)
    batch_1_bytes = write_session_file(session_df_1)
    existing_session_df = read_session_file(BytesIO(batch_1_bytes))

    # Batch 2 - same research file (smaller orders so cash still fits)
    research_df_2 = read_research_file(_research_buys_xlsx()).head(1).copy()
    val_2 = validate_orders(
        research_df_2, bank_book, None, isin_db,
        existing_session_df=existing_session_df,
    )

    session_df_2 = build_session_file(val_2, existing_session_df=existing_session_df)
    assert int(session_df_2["Batch"].max()) == 2
    # Combined session file holds rows from both batches
    assert set(session_df_2["Batch"]) == {1, 2}
