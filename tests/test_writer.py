"""Tests for utils/writer.py - session file + allocation file formatting.

The allocation file's exact formatting (Aptos Narrow, borders, alignment,
number formats) is what Orbis expects on import, so every aspect is locked
down with a test."""
from io import BytesIO

import pandas as pd
import pytest
from openpyxl import load_workbook

from utils.writer import (
    to_excel_bytes, write_session_file, write_allocation_file,
)
from part2.allocator import ALLOCATION_COLUMNS


# ── Helpers ──────────────────────────────────────────────────────────────────

def _open(b: bytes):
    return load_workbook(BytesIO(b))


def _session_df():
    return pd.DataFrame({
        "S.No":   [1, 2, 3],
        "Batch":  [1, 1, 1],
        "OFIN":   ["OF001", "OF002", "OF003"],
        "Client": ["A", "B", "C"],
        "Ticker": ["X", "X", "Y"],
        "ISIN":   ["INE001", "INE001", "INE002"],
        "Direction": ["BUY", "BUY", "SELL"],
        "Qty":    [100, 200, 50],
        "Ref Price": [10.0, 10.0, 20.0],
        "CP Code": ["VALID", "", "VALID2"],  # middle row is blank
    })


def _allocation_df():
    df = pd.DataFrame({
        "S.No": [1, 2],
        "Client Name": ["Client A", "Client B"],
        "CustomerNo": ["OF001", "OF002"],
        "TradeDate": [pd.Timestamp("2026-05-18"), pd.Timestamp("2026-05-18")],
        "Exchange Type": ["NSE", "NSE"],
        "Settlement No": [None, None],
        "ISIN No": ["INE001", "INE001"],
        "Buy/ Sell": ["Buy", "Buy"],
        "Input Quantity": [100, 200],
        "InputBrokerage": [10.0, 20.0],
        "InputSTT": [5.0, 10.0],
        "InputStampDuty": [1.0, 2.0],
        "InputSEBIChrg": [0.05, 0.10],
        "InputTurnOver": [1.5, 3.0],
        "InputOtherCharges": [0.0, 0.0],
        "InputGST": [2.0, 4.0],
        "InputNetAmount": [1000.0, 2000.0],
        "InputNetRate": [10.0, 10.0],
        "CP CODE": ["ORBIS001", "ORBIS002"],
    })
    return df[ALLOCATION_COLUMNS]


# ── Generic writer ───────────────────────────────────────────────────────────

def test_to_excel_bytes_round_trip():
    df = pd.DataFrame({"A": [1, 2], "B": ["x", "y"]})
    out = to_excel_bytes(df, "Test")
    wb = _open(out)
    assert "Test" in wb.sheetnames
    ws = wb["Test"]
    assert ws.cell(1, 1).value == "A"
    assert ws.cell(2, 1).value == 1
    assert ws.cell(3, 2).value == "y"


# ── Session file writer ──────────────────────────────────────────────────────

def test_session_writer_header_is_bold_and_columns_match():
    out = write_session_file(_session_df())
    ws = _open(out)["Session"]
    assert ws.cell(1, 1).font.bold is True
    headers = [ws.cell(1, c).value for c in range(1, 11)]
    assert headers == ["S.No", "Batch", "OFIN", "Client", "Ticker",
                       "ISIN", "Direction", "Qty", "Ref Price", "CP Code"]


def test_session_writer_blank_cp_code_cell_has_amber_fill():
    """Per the spec, blank CP Code cells get amber #FEF3C7 fill so ops can
    spot them before uploading to Part 2."""
    out = write_session_file(_session_df())
    ws = _open(out)["Session"]
    cp_col = 10  # CP Code is column J

    # Row 2 (OF001) has a CP Code - no fill
    fill_row2 = ws.cell(2, cp_col).fill
    fill_is_blank_row2 = (
        fill_row2.fill_type is None
        or (fill_row2.start_color is not None
            and "FEF3C7" not in str(fill_row2.start_color.rgb).upper())
    )
    assert fill_is_blank_row2

    # Row 3 (OF002) has blank CP Code - amber fill
    fill_row3 = ws.cell(3, cp_col).fill
    assert "FEF3C7" in str(fill_row3.start_color.rgb).upper()


# ── Allocation file writer ───────────────────────────────────────────────────

def test_allocation_writer_has_exactly_19_columns():
    out = write_allocation_file(_allocation_df())
    ws = _open(out)["Allocation"]
    assert ws.max_column == 19
    headers = [ws.cell(1, c).value for c in range(1, 20)]
    assert headers == ALLOCATION_COLUMNS


def test_allocation_writer_uses_aptos_narrow_size_11_throughout():
    """Per Orbis spec, every header + data cell is Aptos Narrow size 11."""
    out = write_allocation_file(_allocation_df())
    ws = _open(out)["Allocation"]
    for r in [1, 2, 3]:  # header + both data rows
        for c in [1, 5, 10, 19]:
            cell = ws.cell(r, c)
            assert cell.font.name == "Aptos Narrow", (
                f"Cell ({r},{c}) font is {cell.font.name!r}, expected Aptos Narrow"
            )
            assert cell.font.size == 11


def test_allocation_writer_header_is_bold_no_fill():
    out = write_allocation_file(_allocation_df())
    ws = _open(out)["Allocation"]
    assert ws.cell(1, 1).font.bold is True
    # No fill color (or unset)
    fill = ws.cell(1, 1).fill
    assert fill.fill_type is None or fill.start_color is None or \
           str(fill.start_color.rgb).upper() in ("00000000", "FFFFFFFF", "NONE")


def test_allocation_writer_settlement_no_always_blank():
    """Per spec, Settlement No is always written as blank (None)."""
    out = write_allocation_file(_allocation_df())
    ws = _open(out)["Allocation"]
    col = ALLOCATION_COLUMNS.index("Settlement No") + 1
    for r in range(2, ws.max_row + 1):
        assert ws.cell(r, col).value is None


def test_allocation_writer_charge_columns_have_2dp_number_format():
    out = write_allocation_file(_allocation_df())
    ws = _open(out)["Allocation"]
    for col_name in ["InputBrokerage", "InputSTT", "InputStampDuty",
                     "InputSEBIChrg", "InputOtherCharges", "InputGST",
                     "InputNetAmount", "InputNetRate"]:
        c = ALLOCATION_COLUMNS.index(col_name) + 1
        assert ws.cell(2, c).number_format == "0.00", (
            f"{col_name} number_format is {ws.cell(2, c).number_format!r}"
        )


def test_allocation_writer_input_turnover_displays_2dp():
    """InputTurnOver is stored at 4dp internally but the Excel display format
    is '0.00' so it visually matches the other charge columns."""
    out = write_allocation_file(_allocation_df())
    ws = _open(out)["Allocation"]
    col = ALLOCATION_COLUMNS.index("InputTurnOver") + 1
    assert ws.cell(2, col).number_format == "0.00"


def test_allocation_writer_trade_date_format_is_dd_mm_yyyy():
    out = write_allocation_file(_allocation_df())
    ws = _open(out)["Allocation"]
    col = ALLOCATION_COLUMNS.index("TradeDate") + 1
    assert ws.cell(2, col).number_format == "DD-MM-YYYY"


def test_allocation_writer_all_cells_have_thin_borders():
    out = write_allocation_file(_allocation_df())
    ws = _open(out)["Allocation"]
    # Spot-check header row + both data rows on a few columns
    for r in [1, 2, 3]:
        for c in [1, 5, 10, 19]:
            cell = ws.cell(r, c)
            assert cell.border.left.style  == "thin"
            assert cell.border.right.style == "thin"
            assert cell.border.top.style   == "thin"
            assert cell.border.bottom.style == "thin"


def test_allocation_writer_client_name_left_aligned():
    out = write_allocation_file(_allocation_df())
    ws = _open(out)["Allocation"]
    col = ALLOCATION_COLUMNS.index("Client Name") + 1
    assert ws.cell(2, col).alignment.horizontal == "left"
    assert ws.cell(2, col).alignment.vertical == "center"


def test_allocation_writer_non_client_columns_center_aligned():
    out = write_allocation_file(_allocation_df())
    ws = _open(out)["Allocation"]
    for col_name in ["ISIN No", "Buy/ Sell", "InputBrokerage", "CP CODE"]:
        c = ALLOCATION_COLUMNS.index(col_name) + 1
        assert ws.cell(2, c).alignment.horizontal == "center", (
            f"{col_name} alignment is {ws.cell(2, c).alignment.horizontal!r}"
        )


def test_allocation_writer_buy_sell_value_preserves_title_case():
    """Orbis import expects Buy / Sell, not BUY / SELL. The allocator
    produces title-case; the writer must not normalise it back."""
    out = write_allocation_file(_allocation_df())
    ws = _open(out)["Allocation"]
    col = ALLOCATION_COLUMNS.index("Buy/ Sell") + 1
    assert ws.cell(2, col).value == "Buy"
