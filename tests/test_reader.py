"""Tests for utils/reader.py - Direction normalisation, blank-row drops,
plus edge cases for the bank book and scrip-wise report readers."""
from io import BytesIO

import pytest
import pandas as pd
from openpyxl import Workbook

from utils.reader import read_research_file, read_bank_book, read_scrip_wise_report


def _make_research_file(tmp_path, rows: list[list]) -> str:
    """Write a minimal research xlsx with the standard header + given data rows.
    Returns the path. Header columns mirror what reader.py expects to find."""
    wb = Workbook()
    ws = wb.active
    ws.append(["S.No", "OFIN", "Client", "Ticker",
               "Direction", "Qty", "Ref Price", "Value", "CP Code"])
    for r in rows:
        ws.append(r)
    p = tmp_path / "research.xlsx"
    wb.save(p)
    return str(p)


def test_blank_direction_rows_dropped_silently(tmp_path):
    """Trailing summary / blank-Direction rows are not orders; they should be
    silently dropped during parsing rather than raising 'NONE' as an unknown
    Direction value."""
    path = _make_research_file(tmp_path, [
        [1, "OF001", "A", "STOCK", "Buy",  100, 10.0, 1000.0, "X"],
        [2, "OF002", "B", "STOCK", "Sell",  50, 10.0,  500.0, "X"],
        # Totals row - only Amount filled, Direction blank
        [None, None, None, None,  None,   None, None,  1500.0, None],
    ])
    with open(path, "rb") as f:
        df = read_research_file(f)
    assert len(df) == 2
    assert list(df["Direction"]) == ["BUY", "SELL"]


def test_direction_case_and_whitespace_normalised(tmp_path):
    """Case + whitespace variations of BUY/SELL normalise to canonical form."""
    path = _make_research_file(tmp_path, [
        [1, "OF001", "A", "X", "buy",        5, 10.0,   50.0, "Y"],
        [2, "OF002", "B", "X", "  Sell  ",  15, 10.0,  150.0, "Y"],
        [3, "OF003", "C", "X", "Buy",       10, 10.0,  100.0, "Y"],
        [4, "OF004", "D", "X", "SELL",      20, 10.0,  200.0, "Y"],
    ])
    with open(path, "rb") as f:
        df = read_research_file(f)
    assert list(df["Direction"]) == ["BUY", "SELL", "BUY", "SELL"]


def test_direction_aliases_normalise_to_canonical(tmp_path):
    """Comprehensive alias coverage: every entry in _DIRECTION_ALIASES should
    map to either BUY or SELL after the reader processes it. Locks down the
    contract so accidental removal of an alias is caught by the test suite."""
    buy_inputs = [
        "BUY", "B", "BUYS", "BUYING", "BOUGHT",
        "PURCHASE", "PURCHASES", "PURCHASED", "PURCHASING", "PURCH",
        "LONG", "ADD", "ACQUIRE", "ENTRY", "ENTER",
    ]
    sell_inputs = [
        "SELL", "S", "SELLS", "SELLING", "SOLD",
        "SALE", "SALES",
        "SHORT", "TRIM", "REDUCE", "DISPOSE", "EXIT",
    ]
    # Mix in case variations to confirm case-insensitivity coexists with aliasing
    rows = []
    for i, txt in enumerate(buy_inputs + sell_inputs, start=1):
        # Alternate casing: every other input is lowercase
        d = txt.lower() if i % 2 == 0 else txt
        rows.append([i, f"OF{i:03d}", f"C{i}", "X", d, 10, 10.0, 100.0, "Y"])
    path = _make_research_file(tmp_path, rows)
    with open(path, "rb") as f:
        df = read_research_file(f)
    expected = (["BUY"] * len(buy_inputs)) + (["SELL"] * len(sell_inputs))
    assert list(df["Direction"]) == expected


def test_unrecognised_direction_passes_through_reader(tmp_path):
    """Genuinely unknown Direction values (not in alias map) survive the
    reader so the UI's detection helper can surface them as an error."""
    path = _make_research_file(tmp_path, [
        [1, "OF001", "A", "STOCK", "BUY",     100, 10.0, 1000.0, "X"],
        [2, "OF002", "B", "STOCK", "BUYY",     50, 10.0,  500.0, "X"],  # typo
        [3, "OF003", "C", "STOCK", "FOOBAR",   30, 10.0,  300.0, "X"],
        [4, "OF004", "D", "STOCK", "TRANSFER", 20, 10.0,  200.0, "X"],
    ])
    with open(path, "rb") as f:
        df = read_research_file(f)
    assert len(df) == 4
    # Unknown values survive uppercased - reader does not silently coerce them.
    assert "BUYY" in set(df["Direction"])
    assert "FOOBAR" in set(df["Direction"])
    assert "TRANSFER" in set(df["Direction"])


# ── Bank book reader ────────────────────────────────────────────────────────

def _bank_book_xlsx(rows: list[list], header_at: int = 0,
                    sheet_name: str = "Bank Balance Summary") -> BytesIO:
    """Synthetic bank book file. header_at adds blank rows before the header
    to simulate Orbis exports that have title rows at the top."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for _ in range(header_at):
        ws.append([])
    ws.append(["OFIN Code", "Balance"])
    for r in rows:
        ws.append(r)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def test_bank_book_parses_ofin_to_balance_dict():
    f = _bank_book_xlsx([
        ["OF001", 500000.0],
        ["OF002", 200000.0],
        ["OF003", 100000.0],
    ])
    result = read_bank_book(f)
    assert result == {"OF001": 500000.0, "OF002": 200000.0, "OF003": 100000.0}


def test_bank_book_skips_title_rows_above_header():
    """Orbis exports often have 2-3 title rows above the actual header.
    Reader scans for the row containing 'OFIN Code' rather than hardcoding row 0."""
    f = _bank_book_xlsx(
        [["OF001", 500000.0]],
        header_at=3,
    )
    result = read_bank_book(f)
    assert result == {"OF001": 500000.0}


def test_bank_book_skips_total_rows():
    """Rows containing 'Total' (anywhere in the row, case-insensitive) are
    summary rows and must not contribute to the bank-balance dict."""
    f = _bank_book_xlsx([
        ["OF001", 500000.0],
        ["Total", 500000.0],   # summary row, skip
        ["OF002", 200000.0],
    ])
    result = read_bank_book(f)
    assert "Total" not in result
    assert result == {"OF001": 500000.0, "OF002": 200000.0}


def test_bank_book_preserves_negative_balances():
    """Real-world: ops cash accounts can go negative. Reader must not coerce
    or drop these rows."""
    f = _bank_book_xlsx([
        ["OF001", -5000.0],
        ["OF002", 100000.0],
    ])
    result = read_bank_book(f)
    assert result["OF001"] == -5000.0


def test_bank_book_missing_sheet_raises_clear_error():
    wb = Workbook()
    ws = wb.active
    ws.title = "WrongSheetName"
    ws.append(["OFIN Code", "Balance"])
    ws.append(["OF001", 500000.0])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    with pytest.raises(ValueError, match=r"Bank Balance Summary"):
        read_bank_book(buf)


def test_bank_book_missing_ofin_column_raises_clear_error():
    """If the file has no 'OFIN Code' header at all, ValueError with helpful message."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Bank Balance Summary"
    ws.append(["WrongHeader1", "WrongHeader2"])
    ws.append(["OF001", 500000.0])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    with pytest.raises(ValueError, match=r"OFIN Code"):
        read_bank_book(buf)


# ── Scrip-wise reader ───────────────────────────────────────────────────────

def _scrip_wise_xlsx(rows: list[list], sheet_name: str = "ClientA",
                     header_at: int = 0) -> BytesIO:
    """Synthetic scrip-wise report. Sheet name varies per file (named after
    the first client), so reader uses sheet index 0 rather than name."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for _ in range(header_at):
        ws.append([])
    ws.append(["Scrip Name", "Item No", "Client Code", "Quantity"])
    for r in rows:
        ws.append(r)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def test_scrip_wise_parses_holdings_dataframe():
    f = _scrip_wise_xlsx([
        ["HDFCBANK",  "INE040A01034", "OF001", 100],
        ["KOTAKBANK", "INE237A01036", "OF001",  50],
    ])
    df = read_scrip_wise_report(f)
    assert len(df) == 2
    assert set(df.columns) >= {"OFIN", "Scrip Name", "ISIN", "Quantity"}
    assert df[df["OFIN"] == "OF001"]["Scrip Name"].tolist() == ["HDFCBANK", "KOTAKBANK"]


def test_scrip_wise_skips_scrip_total_summary_rows():
    """'Scrip Total' rows are aggregations across clients - drop them."""
    f = _scrip_wise_xlsx([
        ["HDFCBANK",    "INE040A01034", "OF001", 100],
        ["Scrip Total", "",             "",      100],  # summary
        ["KOTAKBANK",   "INE237A01036", "OF002",  50],
    ])
    df = read_scrip_wise_report(f)
    assert len(df) == 2
    assert "Scrip Total" not in df["Scrip Name"].values


def test_scrip_wise_skips_blank_rows():
    f = _scrip_wise_xlsx([
        ["HDFCBANK",  "INE040A01034", "OF001", 100],
        ["",          "",             "",      ""],   # blank
        [None,        None,           None,    None],  # None
        ["KOTAKBANK", "INE237A01036", "OF002",  50],
    ])
    df = read_scrip_wise_report(f)
    assert len(df) == 2


def test_scrip_wise_uses_first_sheet_by_index_not_name():
    """Sheet name changes daily (named after the first client). Code must
    read from sheet index 0, not by matching a name."""
    f = _scrip_wise_xlsx(
        [["HDFCBANK", "INE040A01034", "OF001", 100]],
        sheet_name="RandomSheetNameThatChangesDaily",
    )
    df = read_scrip_wise_report(f)
    assert len(df) == 1


def test_scrip_wise_skips_title_rows_above_header():
    """Header isn't always at row 0; reader scans for 'Scrip Name'."""
    f = _scrip_wise_xlsx(
        [["HDFCBANK", "INE040A01034", "OF001", 100]],
        header_at=4,
    )
    df = read_scrip_wise_report(f)
    assert len(df) == 1
    assert df.iloc[0]["Scrip Name"] == "HDFCBANK"


def test_scrip_wise_missing_required_column_raises_value_error():
    wb = Workbook()
    ws = wb.active
    ws.title = "Anything"
    ws.append(["Scrip Name", "Item No"])  # missing Client Code + Quantity
    ws.append(["HDFCBANK", "INE040A01034"])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    with pytest.raises(ValueError):
        read_scrip_wise_report(buf)


def test_scrip_wise_ofin_coerced_to_string_and_stripped():
    """OFIN sometimes arrives with whitespace or as a numeric-looking string.
    Reader must always produce a clean string for downstream matching."""
    f = _scrip_wise_xlsx([
        ["HDFCBANK",  "INE040A01034", "  OF001  ",  100],
        ["KOTAKBANK", "INE237A01036", "00012345",    50],
    ])
    df = read_scrip_wise_report(f)
    assert "OF001" in df["OFIN"].values   # stripped
    assert "00012345" in df["OFIN"].values  # leading zeros preserved as string


# ── Session-file round-trip: blank CP Code preservation ────────────────────

def test_session_file_blank_cp_code_round_trips_as_empty_string(tmp_path):
    """Regression: blank Excel cells came back as the literal string 'nan'
    under dtype=str, which is truthy - the allocator's InCred CP-code
    fallback then skipped the fallback. Reader must return '' for blanks."""
    from utils.writer import write_session_file
    from utils.reader import read_session_file
    session_df = pd.DataFrame({
        "S.No":   [1, 2],
        "Batch":  [1, 1],
        "OFIN":   ["OF001", "OF002"],
        "Client": ["A", "B"],
        "Ticker": ["X", "Y"],
        "ISIN":   ["INE001", "INE002"],
        "Direction": ["BUY", "SELL"],
        "Qty":    [100, 50],
        "Ref Price": [10.0, 20.0],
        "CP Code": ["VALID", ""],  # second row blank
    })
    bytes_ = write_session_file(session_df)
    round_trip = read_session_file(BytesIO(bytes_))
    cps = list(round_trip["CP Code"])
    assert cps[0] == "VALID"
    assert cps[1] == "", f"Blank CP Code became {cps[1]!r} after round-trip"
